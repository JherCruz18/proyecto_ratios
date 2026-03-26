import pandas as pd
from sqlalchemy import text
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import IconSet, FormatObject
from db import engine

# ==============================
# METAS POR SUCURSAL (decimales reales)
# ==============================
METAS = {
    1: 0.0042,   # CERRO
    3: 0.0065,   # ACANTILADO / LARCOMAR
    4: 0.0050,   # PRADO
    5: 0.0050    # REMANSO
}


def exportar_carbon_excel(
    id_sucursal,
    nombre_sucursal,
    fecha_inicio,
    fecha_fin,
    plantilla="formato.xlsx",
    salida="Reporte_Carbon.xlsx"
):
    # ==============================
    # CONSULTA SQL
    # ==============================
    with engine.connect() as conn:
        df = pd.read_sql(
            text("""
                SELECT
                    ri.fecha,
                    ri.stock_inicial,
                    ri.ingreso,
                    COALESCE(ri.reposicion, 0) AS reposicion,
                    ri.stock_final,
                    ri.venta_total,
                    i.nombre AS tipo_carbon
                FROM registro_insumo ri
                JOIN insumos i ON i.id_insumo = ri.id_insumo
                WHERE ri.id_sucursal = :id_sucursal
                  AND ri.id_insumo = 1
                  AND ri.estado = 1
                  AND ri.fecha BETWEEN :fi AND :ff
                ORDER BY ri.fecha ASC
            """),
            conn,
            params={
                "id_sucursal": id_sucursal,
                "fi": fecha_inicio,
                "ff": fecha_fin
            }
        )

    if df.empty:
        raise ValueError("No hay datos para el rango seleccionado")

    # ==============================
    # CARGAR PLANTILLA
    # ==============================
    wb = load_workbook(plantilla)
    ws = wb.active

    ws["D3"] = nombre_sucursal

    # ==============================
    # ESCRIBIR DATOS
    # ==============================
    fila = 7
    for _, r in df.iterrows():
        ws[f"C{fila}"] = r["fecha"].strftime("%d/%m/%Y")
        ws[f"D{fila}"] = r["stock_inicial"]
        ws[f"E{fila}"] = r["ingreso"]
        ws[f"G{fila}"] = r["reposicion"]
        ws[f"H{fila}"] = r["stock_final"]
        ws[f"I{fila}"] = r["venta_total"]
        ws[f"K{fila}"] = r["tipo_carbon"]
        fila += 1

    ultima_fila = fila - 1

    # ==============================
    # FORMATO CONDICIONAL (SEMÁFORO FINAL)
    # ==============================

    # Meta dinámica por sucursal
    meta = METAS.get(id_sucursal, 0.0042)

    # REGLA:
    # 🔴 Rojo     → valor > meta
    # 🟢 Verde    → valor >= 0 y <= meta
    # 🟡 Amarillo → valor negativo

    fo_rojo = FormatObject(type="num", val=str(meta))  # rojo
    fo_amarillo = FormatObject(type="num", val="0")    # amarillo si negativo
    fo_verde = FormatObject(type="num", val="-1")      # define verde para todo >= 0

    iconset = IconSet(
        iconSet="3TrafficLights2",   # iconos originales de Excel
        reverse=False,
        showValue=True,
        cfvo=[fo_rojo, fo_amarillo, fo_verde]
    )

    regla = Rule(type="iconSet", iconSet=iconset)

    # Columna J dinámica
    ws.conditional_formatting.add(f"J7:J{ultima_fila}", regla)

    # VALIDACIÓN EXTRA → J38 (como pediste)
    ws.conditional_formatting.add("J38", regla)

    # ==============================
    # FORZAR CÁLCULO EN EXCEL
    # ==============================
    wb.properties.fullCalcOnLoad = True

    # ==============================
    # GUARDAR ARCHIVO
    # ==============================
    wb.save(salida)

    return salida
